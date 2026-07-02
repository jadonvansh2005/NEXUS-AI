from typing import List

from openpyxl import load_workbook

from rag.loaders.base_loader import (
    BaseLoader
)

from rag.loaders.loader_constants import (
    EXCEL_EXTENSIONS,
    DocumentType
)

from rag.loaders.loader_factory import (
    LoaderFactory
)

from rag.loaders.loader_schema import (
    LoadedDocument
)

from rag.loaders.loader_utils import (
    LoaderUtils
)


class ExcelLoader(

    BaseLoader

):

    SUPPORTED_EXTENSIONS = (

        EXCEL_EXTENSIONS

    )

    def load(

        self

    ) -> LoadedDocument:

        workbook = (

            load_workbook(

                filename=self.absolute_path,

                data_only=True

            )

        )

        sheets: List[str] = []

        total_rows = 0

        total_columns = 0

        for worksheet in workbook.worksheets:

            rows = []

            total_columns = max(

                total_columns,

                worksheet.max_column

            )

            for row in worksheet.iter_rows(

                values_only=True

            ):

                values = [

                    str(cell)

                    for cell in row

                    if cell is not None

                ]

                if values:

                    rows.append(

                        " | ".join(

                            values

                        )

                    )

                    total_rows += 1

            sheet_text = (

                f"Sheet: {worksheet.title}\n"

                + "\n".join(

                    rows

                )

            )

            sheets.append(

                LoaderUtils.normalize_text(

                    sheet_text

                )

            )

        workbook.close()

        full_text = (

            "\n\n".join(

                sheets

            )

        )

        return LoadedDocument(

            text=full_text,

            document_type=(

                DocumentType.XLSX.value

            ),

            source_path=(

                self.absolute_path

            ),

            filename=(

                self.filename

            ),

            metadata=self.metadata(),

            page_count=len(

                sheets

            ),

            pages=sheets,

            checksum=(

                LoaderUtils.calculate_checksum(

                    self.absolute_path

                )

            ),

            file_size=(

                self.filesize

            ),

            extra={

                "sheet_count": len(

                    workbook.sheetnames

                ),

                "sheet_names": workbook.sheetnames,

                "row_count": total_rows,

                "column_count": total_columns

            }

        )


LoaderFactory.register(

    ExcelLoader

)